import csv
import json
import os
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = Path(os.environ.get("TEMP", ".")) / "cofrinho-catalogo-fontes"
SOURCE = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else DEFAULT_SOURCE.resolve()
OUTPUT = ROOT / "data" / "catalogo-fontes"
ACCESSED = "2026-07-12"


def write_json(name, value):
    OUTPUT.mkdir(parents=True, exist_ok=True)
    (OUTPUT / name).write_text(
        json.dumps(value, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def clean(value):
    return " ".join(str(value or "").strip().split())


def folded(value):
    return "".join(
        char
        for char in unicodedata.normalize("NFD", clean(value)).casefold()
        if unicodedata.category(char) != "Mn"
    )


def read_cbo():
    path = SOURCE / "cbo.csv"
    raw = path.read_bytes()
    text = None
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise RuntimeError("Nao foi possivel decodificar cbo.csv")
    rows = []
    for row in csv.DictReader(text.splitlines(), delimiter=";"):
        code = clean(row.get("CODIGO"))
        title = clean(row.get("TITULO"))
        if code and title:
            rows.append({"codigo": code, "titulo": title})
    return rows


def read_municipalities():
    data = json.loads((SOURCE / "municipios.json").read_text(encoding="utf-8-sig"))
    rows = []
    for item in data:
        if item.get("microrregiao"):
            state = item["microrregiao"]["mesorregiao"]["UF"]
        else:
            state = item["regiao-imediata"]["regiao-intermediaria"]["UF"]
        rows.append(
            {
                "codigo_ibge": str(item["id"]),
                "nome": clean(item["nome"]),
                "uf": state["sigla"],
                "regiao": folded(state["regiao"]["nome"]).replace(" ", "-"),
            }
        )
    return rows


def read_workbook(filename, name_column, code_column, exclusions, contains_exclusions=()):
    workbook = load_workbook(SOURCE / filename, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = []
    seen = set()
    excluded = {folded(value) for value in exclusions}
    contains = tuple(folded(value) for value in contains_exclusions)
    for row in sheet.iter_rows(min_row=7, values_only=True):
        name = clean(row[name_column]) if len(row) > name_column else ""
        code = clean(row[code_column]) if len(row) > code_column else ""
        key = folded(name)
        if not name or key in seen or key in excluded or any(part in key for part in contains):
            continue
        seen.add(key)
        rows.append({"codigo_ibge": code or None, "nome_oficial": name})
    return rows


def main():
    missing = [name for name in ("cbo.csv", "municipios.json", "etnias.xlsx", "linguas.xlsx") if not (SOURCE / name).exists()]
    if missing:
        raise SystemExit(f"Fontes ausentes em {SOURCE}: {', '.join(missing)}")

    occupations = read_cbo()
    municipalities = read_municipalities()
    peoples = read_workbook(
        "etnias.xlsx",
        4,
        5,
        (
            "Outra etnia das Américas",
            "Não determinada",
            "Não sabe",
            "Sem declaração",
            "Mal definida",
        ),
    )
    languages = read_workbook(
        "linguas.xlsx",
        5,
        6,
        (
            "Outra língua das Américas",
            "Não determinada",
            "Não sabe",
            "Sem declaração",
            "Mal definida",
        ),
        ("não especificado",),
    )

    if len(peoples) != 391:
        raise SystemExit(f"Esperados 391 povos/grupos; encontrados {len(peoples)}")
    if len(languages) != 295:
        raise SystemExit(f"Esperadas 295 linguas; encontradas {len(languages)}")
    if len(municipalities) < 5570:
        raise SystemExit(f"Base de municipios incompleta: {len(municipalities)}")
    if len(occupations) < 600:
        raise SystemExit(f"Base CBO incompleta: {len(occupations)}")

    source_common = {
        "data_consulta": ACCESSED,
        "status": "importado_de_fonte_oficial",
    }
    write_json(
        "mte-cbo-ocupacoes.json",
        {
            "schema_version": "1.0.0",
            "fonte": {
                **source_common,
                "titulo": "Classificacao Brasileira de Ocupacoes - CBO 2002, ocupacoes",
                "instituicao": "Ministerio do Trabalho e Emprego",
                "url": "https://www.gov.br/trabalho-e-emprego/pt-br/assuntos/cbo/servicos/downloads",
            },
            "ocupacoes": occupations,
        },
    )
    write_json(
        "ibge-municipios.json",
        {
            "schema_version": "1.0.0",
            "fonte": {
                **source_common,
                "titulo": "API de localidades - municipios",
                "instituicao": "IBGE",
                "url": "https://servicodados.ibge.gov.br/api/v1/localidades/municipios",
            },
            "municipios": municipalities,
        },
    )
    write_json(
        "ibge-povos-indigenas.json",
        {
            "schema_version": "1.0.0",
            "fonte": {
                **source_common,
                "titulo": "Censo Demografico 2022 - Etnias e linguas indigenas, Apendice 1",
                "instituicao": "IBGE",
                "url": "https://ftp.ibge.gov.br/Censos/Censo_Demografico_2022/Etnias_e_Linguas_Indigenas_principais_caracteristicas_sociodemograficas_Resultados_do_universo/Apendices/xlsx/",
            },
            "criterio": "391 nomes oficiais apos retirar categorias residuais e sem declaracao do apendice.",
            "povos": peoples,
        },
    )
    write_json(
        "ibge-linguas-indigenas.json",
        {
            "schema_version": "1.0.0",
            "fonte": {
                **source_common,
                "titulo": "Censo Demografico 2022 - Etnias e linguas indigenas, Apendice 2",
                "instituicao": "IBGE",
                "url": "https://ftp.ibge.gov.br/Censos/Censo_Demografico_2022/Etnias_e_Linguas_Indigenas_principais_caracteristicas_sociodemograficas_Resultados_do_universo/Apendices/xlsx/",
            },
            "criterio": "295 nomes oficiais apos retirar categorias residuais, nao especificadas e sem declaracao do apendice.",
            "linguas": languages,
        },
    )
    print(json.dumps({
        "fonte": str(SOURCE),
        "ocupacoes_cbo": len(occupations),
        "municipios": len(municipalities),
        "povos_indigenas": len(peoples),
        "linguas_indigenas": len(languages),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
